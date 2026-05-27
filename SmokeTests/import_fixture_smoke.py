#!/usr/bin/python3
import json
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "Resources" / "Scripts" / "excel_sync.py"
PYTHONPATH = ROOT / "Resources" / "python"

sys.path.insert(0, str(PYTHONPATH))
from openpyxl import Workbook, load_workbook  # noqa: E402


def run(command, workbook, payload):
    env = os.environ.copy()
    env["PYTHONPATH"] = str(PYTHONPATH)
    result = subprocess.run(
        [sys.executable, str(SCRIPT), command, str(workbook)],
        input=json.dumps(payload),
        text=True,
        capture_output=True,
        env=env,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr or result.stdout)
    decoded = json.loads(result.stdout)
    if not decoded.get("success"):
        raise RuntimeError(decoded.get("error") or result.stdout)
    return decoded


def make_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(["Item Type", "Description", "Manufacturer", "Part Number", "Purchase Date", "Vendor", "Unit Cost", "Quantity", "Total Cost", "Qty Received", "PO Number", "Remaining Inventory", "Notes"])
    ws.append(["Laptop", "Example Laptop", "Example Manufacturer", "LAP-1", "2026-01-01", "Example Vendor", 1000, 5, "=G2*H2", "5/5", "PO-1", 5, "Seed"])
    ws2 = wb.create_sheet("OpEx")
    ws2.append(["Item Type", "Description", "Manufacturer", "Part Number", "Purchase Date", "Vendor", "Unit Cost", "Quantity", "Total Cost", "Qty Received", "PO Number", "Remaining Inventory", "Notes"])
    ws3 = wb.create_sheet("Items Deployed")
    ws3.append(["Item Type", "Description", "Manufacturer", "Part Number", "Qty Deployed", "Deployed To", "Deployed By", "Deployed Date", "Location", "Notes"])
    wb.save(path)


def main():
    tmp = Path(tempfile.mkdtemp(prefix="inventory-import-fixture-"))
    try:
        workbook = tmp / "fixture.xlsx"
        make_workbook(workbook)
        run("append-inventory", workbook, {"items": [{
            "itemType": "Peripheral",
            "description": "Example Dock",
            "manufacturer": "Example Manufacturer",
            "partNumber": "DOCK-1",
            "purchaseDate": "2026-02-01",
            "vendor": "Example Vendor",
            "unitCost": 199,
            "quantity": 3,
            "qtyReceived": 3,
            "poNumber": "PO-2",
            "budgetType": "OpEx",
            "stockroomName": "Main",
            "notes": "Fixture"
        }]})
        run("append-deployed", workbook, {"deployments": [{
            "itemType": "Laptop",
            "description": "Example Laptop",
            "manufacturer": "Example Manufacturer",
            "partNumber": "LAP-1",
            "qtyDeployed": 2,
            "deployedTo": "Example Team",
            "deployedBy": "Smoke Test",
            "deployedDate": "2026-03-01",
            "deployedLocation": "HQ",
            "notes": "Fixture deployment"
        }, {
            "itemType": "Peripheral",
            "description": "Example Dock",
            "manufacturer": "Example Manufacturer",
            "partNumber": "DOCK-1",
            "qtyDeployed": 1,
            "deployedTo": "Example Desk",
            "deployedBy": "Smoke Test",
            "deployedDate": "2026-03-02",
            "deployedLocation": "HQ",
            "notes": "Linked fixture deployment"
        }]})
        run("update-remaining", workbook, {"items": [{
            "partNumber": "LAP-1",
            "poNumber": "PO-1",
            "budgetType": "Capital",
            "remaining": 3
        }]})
        remaining_value = load_workbook(workbook)["Inventory"].cell(row=2, column=12).value
        if remaining_value != 3:
            raise AssertionError(f"update-remaining left remaining inventory at {remaining_value!r}")
        wb = load_workbook(workbook)
        ws = wb["Inventory"]
        ws.append(["Accessory", "Ambiguous Item", "Example Manufacturer", "AMB-1", "2026-04-01", "Example Vendor", 100, 5, "=G3*H3", "5/5", "", 5, "first ambiguous"])
        ws.append(["Accessory", "Ambiguous Item", "Example Manufacturer", "AMB-1", "2026-04-01", "Example Vendor", 200, 9, "=G4*H4", "9/9", "", 9, "second ambiguous"])
        wb.save(workbook)
        run("update-inventory", workbook, {
            "original": {
                "itemType": "Accessory",
                "description": "Ambiguous Item",
                "manufacturer": "Example Manufacturer",
                "partNumber": "AMB-1",
                "purchaseDate": "2026-04-01",
                "vendor": "Example Vendor",
                "unitCost": 200,
                "quantity": 9,
                "qtyReceived": 9,
                "poNumber": "",
                "budgetType": "Capital",
                "notes": "second ambiguous",
            },
            "updated": {
                "itemType": "Accessory",
                "description": "Ambiguous Item",
                "manufacturer": "Example Manufacturer",
                "partNumber": "AMB-1",
                "purchaseDate": "2026-04-01",
                "vendor": "Example Vendor",
                "unitCost": 200,
                "quantity": 9,
                "qtyReceived": 9,
                "poNumber": "",
                "budgetType": "Capital",
                "notes": "second ambiguous updated",
            },
        })
        wb = load_workbook(workbook)
        ws = wb["Inventory"]
        ambiguous_rows = [
            (row, ws.cell(row=row, column=7).value, ws.cell(row=row, column=13).value)
            for row in range(2, ws.max_row + 1)
            if ws.cell(row=row, column=4).value == "AMB-1"
        ]
        if (100, "first ambiguous") not in [(cost, note) for _, cost, note in ambiguous_rows] or (200, "second ambiguous updated") not in [(cost, note) for _, cost, note in ambiguous_rows]:
            raise AssertionError("update-inventory matched the wrong no-PO workbook row")
        run("delete-inventory", workbook, {"items": [{
            "itemType": "Accessory",
            "description": "Ambiguous Item",
            "manufacturer": "Example Manufacturer",
            "partNumber": "AMB-1",
            "purchaseDate": "2026-04-01",
            "vendor": "Example Vendor",
            "unitCost": 200,
            "quantity": 9,
            "qtyReceived": 9,
            "poNumber": "",
            "budgetType": "Capital",
            "notes": "second ambiguous updated",
        }]})
        wb = load_workbook(workbook)
        ws = wb["Inventory"]
        ambiguous_rows = [
            (ws.cell(row=row, column=7).value, ws.cell(row=row, column=13).value)
            for row in range(2, ws.max_row + 1)
            if ws.cell(row=row, column=4).value == "AMB-1"
        ]
        if ambiguous_rows != [(100, "first ambiguous")]:
            raise AssertionError("delete-inventory removed the wrong no-PO workbook row")
        deployed_ws = wb["Items Deployed"]
        deployed_ws.append(["Laptop", "Placeholder zero quantity", "Example Manufacturer", "ZERO-DEP", 0, "Example Team", "Smoke Test", "2026-03-03", "HQ", "should be ignored"])
        wb.save(workbook)
        inventory = run("read-inventory", workbook, {})
        deployed = run("read-deployed", workbook, {})
        if len(inventory.get("items", [])) < 2 or len(deployed.get("deployments", [])) < 1:
            raise AssertionError("fixture workbook did not round-trip expected rows")
        if any(deployment.get("partNumber") == "ZERO-DEP" for deployment in deployed.get("deployments", [])):
            raise AssertionError("read-deployed should ignore zero-quantity placeholder rows")
        run("delete-deployed", workbook, {"deployments": [{
            "itemType": "Laptop",
            "description": "Example Laptop",
            "manufacturer": "Example Manufacturer",
            "partNumber": "LAP-1",
            "qtyDeployed": 2,
            "deployedTo": "Example Team",
            "deployedBy": "Smoke Test",
            "deployedDate": "2026-03-01",
            "deployedLocation": "HQ",
            "notes": "Fixture deployment"
        }]})
        run("delete-inventory", workbook, {"items": [{
            "itemType": "Peripheral",
            "description": "Example Dock",
            "manufacturer": "Example Manufacturer",
            "partNumber": "DOCK-1",
            "purchaseDate": "2026-02-01",
            "vendor": "Example Vendor",
            "unitCost": 199,
            "quantity": 3,
            "qtyReceived": 3,
            "poNumber": "PO-2",
            "budgetType": "OpEx",
            "notes": "Fixture"
        }], "deployments": [{
            "itemType": "Peripheral",
            "description": "Example Dock",
            "manufacturer": "Example Manufacturer",
            "partNumber": "DOCK-1",
            "qtyDeployed": 1,
            "deployedTo": "Example Desk",
            "deployedBy": "Smoke Test",
            "deployedDate": "2026-03-02",
            "deployedLocation": "HQ",
            "notes": "Linked fixture deployment"
        }]})
        inventory_after_delete = run("read-inventory", workbook, {})
        deployed_after_delete = run("read-deployed", workbook, {})
        if any(item.get("partNumber") == "DOCK-1" for item in inventory_after_delete.get("items", [])):
            raise AssertionError("delete-inventory did not remove the matching workbook row")
        if deployed_after_delete.get("deployments", []):
            raise AssertionError("delete-deployed did not remove the matching workbook row")
        print("import_fixture_smoke=ok")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    main()
