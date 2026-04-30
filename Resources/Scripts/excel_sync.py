#!/usr/bin/env python3
"""
excel_sync.py - Helper script to modify an inventory workbook using openpyxl.

Usage:
    echo '<json>' | python3 excel_sync.py <command> <excel_path>

Commands:
    append-inventory   - Append rows to the Inventory or OpEx sheet
    append-deployed    - Append rows to the Items Deployed sheet
    update-inventory   - Update an existing row in Inventory or OpEx
    delete-inventory   - Delete existing rows from Inventory or OpEx
    delete-deployed    - Delete existing rows from Items Deployed
    update-remaining   - Update REMAINING INVENTORY (col L) by part number
"""

import json
import sys
from copy import copy
from datetime import date, datetime

try:
    import openpyxl
    from openpyxl.formula.translate import Translator
except ImportError:
    print(json.dumps({"success": False, "error": "openpyxl is not installed. Run: pip3 install openpyxl"}))
    sys.exit(1)


def parse_date(value):
    """Normalize supported date inputs for openpyxl writes."""
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None

        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"):
            try:
                return datetime.strptime(cleaned, fmt)
            except ValueError:
                continue

        return cleaned

    return value


def parse_int(value, default=0):
    """Safely coerce spreadsheet values to integers."""
    if value in (None, ""):
        return default

    if isinstance(value, bool):
        return int(value)

    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return default


def parse_float(value, default=0):
    """Safely coerce spreadsheet values to floats."""
    if value in (None, ""):
        return default

    try:
        return float(str(value).replace(",", "").strip())
    except (ValueError, TypeError, AttributeError):
        return default


def has_row_content(ws, row, columns):
    """Return True when any inspected column in the row contains meaningful content."""
    for col in columns:
        value = ws.cell(row=row, column=col).value
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return True
    return False


def find_last_data_row(ws):
    """Find the last row that contains any data."""
    last_row = 1  # at minimum, row 1 is headers
    for row in range(ws.max_row, 0, -1):
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=row, column=col).value is not None:
                return row
        # If the entire row is empty, keep searching upward
    return last_row


def clone_row_format(ws, source_row, target_row, max_column):
    """Copy row styling and formulas so appended rows inherit workbook formatting."""
    if source_row < 1 or target_row < 1 or source_row == target_row:
        return

    source_dimensions = ws.row_dimensions[source_row]
    target_dimensions = ws.row_dimensions[target_row]
    if source_dimensions.height is not None:
        target_dimensions.height = source_dimensions.height
    target_dimensions.hidden = source_dimensions.hidden
    target_dimensions.outlineLevel = source_dimensions.outlineLevel

    for col in range(1, max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)

        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format

        if source_cell.data_type == "f" and source_cell.value:
            try:
                target_cell.value = Translator(str(source_cell.value), origin=source_cell.coordinate).translate_formula(target_cell.coordinate)
            except Exception:
                target_cell.value = source_cell.value


def append_inventory(wb, data, filepath):
    """Append new rows to the Inventory or OpEx sheet."""
    sheet_name = data.get("sheet", "Inventory")
    items = data.get("items", [])

    if not items:
        return {"success": True, "rowsAdded": 0}

    if sheet_name not in wb.sheetnames:
        return {"success": False, "error": f"Sheet '{sheet_name}' not found in workbook. Available sheets: {wb.sheetnames}"}

    ws = wb[sheet_name]
    last_row = find_last_data_row(ws)
    rows_added = 0

    for item in items:
        row = last_row + 1 + rows_added
        clone_row_format(ws, last_row, row, 13)

        ws.cell(row=row, column=1, value=item.get("itemType", ""))
        ws.cell(row=row, column=2, value=item.get("description", ""))
        ws.cell(row=row, column=3, value=item.get("manufacturer", ""))
        ws.cell(row=row, column=4, value=item.get("partNumber", ""))

        # Purchase date as datetime
        purchase_date = parse_date(item.get("purchaseDate", ""))
        cell_e = ws.cell(row=row, column=5, value=purchase_date)
        if isinstance(purchase_date, datetime):
            cell_e.number_format = "MM/DD/YYYY"

        ws.cell(row=row, column=6, value=item.get("vendor", ""))

        # Cost per item as number
        unit_cost = item.get("unitCost")
        if unit_cost not in (None, ""):
            cell_g = ws.cell(row=row, column=7, value=parse_float(unit_cost))
            cell_g.number_format = '#,##0.00'

        # Purchase quantity as number
        quantity = item.get("quantity")
        if quantity not in (None, ""):
            ws.cell(row=row, column=8, value=parse_int(quantity))

        # Total cost as formula
        cell_i = ws.cell(row=row, column=9)
        if not cell_i.value:
            cell_i.value = f"=G{row}*H{row}"

        # Qty received as workbook-style progress string, e.g. 0/5 or 5/5
        qty_received = parse_int(item.get("qtyReceived", 0))
        quantity_int = parse_int(item.get("quantity", 0))
        if quantity_int > 0:
            ws.cell(row=row, column=10, value=f"{qty_received}/{quantity_int}")
        else:
            ws.cell(row=row, column=10, value=str(qty_received) if qty_received else "")

        # PO number stored as string to preserve leading zeros
        po_number = item.get("poNumber", "")
        cell_k = ws.cell(row=row, column=11, value=str(po_number) if po_number else "")

        # Remaining inventory as number
        remaining = item.get("remainingInventory")
        if remaining not in (None, ""):
            ws.cell(row=row, column=12, value=parse_int(remaining))

        ws.cell(row=row, column=13, value=item.get("notes", ""))

        rows_added += 1

    wb.save(filepath)
    return {"success": True, "rowsAdded": rows_added}


def normalized_text(value):
    return str(value or "").strip().casefold()


def normalized_date(value):
    parsed = parse_date(value)
    if isinstance(parsed, datetime):
        return parsed.strftime("%m/%d/%Y")
    if parsed is None:
        return ""
    return str(parsed).strip().casefold()


def normalized_quantity(value):
    parsed = parse_int(value, default=None)
    return "" if parsed is None else str(parsed)


def normalized_money(value):
    parsed = parse_float(value, default=None)
    return "" if parsed is None else f"{parsed:.2f}"


def normalized_qty_received(value):
    if isinstance(value, str) and "/" in value:
        value = value.split("/", 1)[0]
    return normalized_quantity(value)


def sheet_for_budget(budget_type):
    return "OpEx" if normalized_text(budget_type) == "opex" else "Inventory"


def inventory_match_score(ws, row, item):
    score = 0

    checks = [
        (1, normalized_text(item.get("itemType", ""))),
        (2, normalized_text(item.get("description", ""))),
        (3, normalized_text(item.get("manufacturer", ""))),
        (4, normalized_text(item.get("partNumber", ""))),
        (5, normalized_date(item.get("purchaseDate", ""))),
        (6, normalized_text(item.get("vendor", ""))),
        (7, normalized_money(item.get("unitCost", ""))),
        (8, normalized_quantity(item.get("quantity", ""))),
        (10, normalized_qty_received(item.get("qtyReceived", ""))),
        (11, normalized_text(item.get("poNumber", ""))),
        (13, normalized_text(item.get("notes", ""))),
    ]

    for column, expected in checks:
        actual_value = ws.cell(row=row, column=column).value
        if column == 5:
            actual = normalized_date(actual_value)
        elif column == 7:
            actual = normalized_money(actual_value)
        elif column == 8:
            actual = normalized_quantity(actual_value)
        elif column == 10:
            actual = normalized_qty_received(actual_value)
        else:
            actual = normalized_text(actual_value)
        if expected and actual == expected:
            score += 1

    return score


def find_inventory_row(wb, item, minimum_score=6, early_return_score=8, used_rows=None):
    if used_rows is None:
        used_rows = set()
    preferred_sheet = sheet_for_budget(item.get("budgetType", "Capital"))
    search_order = [preferred_sheet] + [name for name in ["Inventory", "OpEx"] if name != preferred_sheet]

    best_match = None
    best_score = 0

    for sheet_name in search_order:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        last_row = find_last_data_row(ws)

        for row in range(2, last_row + 1):
            if (sheet_name, row) in used_rows:
                continue
            if not has_row_content(ws, row, [1, 2, 3, 4, 6, 11]):
                continue
            score = inventory_match_score(ws, row, item)
            if score > best_score:
                best_match = (sheet_name, row)
                best_score = score
            if score >= early_return_score:
                return (sheet_name, row)

    return best_match if best_score >= minimum_score else None


def write_inventory_row(ws, row, item):
    ws.cell(row=row, column=1, value=item.get("itemType", ""))
    ws.cell(row=row, column=2, value=item.get("description", ""))
    ws.cell(row=row, column=3, value=item.get("manufacturer", ""))
    ws.cell(row=row, column=4, value=item.get("partNumber", ""))

    purchase_date = parse_date(item.get("purchaseDate", ""))
    cell_e = ws.cell(row=row, column=5, value=purchase_date)
    if isinstance(purchase_date, datetime):
        cell_e.number_format = "MM/DD/YYYY"

    ws.cell(row=row, column=6, value=item.get("vendor", ""))

    unit_cost = item.get("unitCost")
    cell_g = ws.cell(row=row, column=7, value=parse_float(unit_cost))
    cell_g.number_format = "#,##0.00"

    quantity_int = parse_int(item.get("quantity", 0))
    ws.cell(row=row, column=8, value=quantity_int)

    cell_i = ws.cell(row=row, column=9)
    if not cell_i.value:
        cell_i.value = f"=G{row}*H{row}"

    qty_received = parse_int(item.get("qtyReceived", 0))
    ws.cell(row=row, column=10, value=f"{qty_received}/{quantity_int}" if quantity_int > 0 else str(qty_received) if qty_received else "")

    po_number = item.get("poNumber", "")
    ws.cell(row=row, column=11, value=str(po_number) if po_number else "")

    remaining = item.get("remainingInventory")
    if remaining not in (None, ""):
        ws.cell(row=row, column=12, value=parse_int(remaining))

    ws.cell(row=row, column=13, value=item.get("notes", ""))


def update_inventory(wb, data, filepath):
    original = data.get("original", {})
    updated = data.get("updated", {})
    if not original or not updated:
        return {"success": False, "error": "Both original and updated inventory payloads are required."}

    match = find_inventory_row(wb, original)
    if not match:
        return {"success": False, "error": "Could not locate the original inventory row in Excel."}

    current_sheet_name, current_row = match
    target_sheet_name = sheet_for_budget(updated.get("budgetType", "Capital"))

    if target_sheet_name not in wb.sheetnames:
        return {"success": False, "error": f"Sheet '{target_sheet_name}' not found in workbook."}

    if current_sheet_name == target_sheet_name:
        ws = wb[current_sheet_name]
        write_inventory_row(ws, current_row, updated)
    else:
        source_ws = wb[current_sheet_name]
        target_ws = wb[target_sheet_name]
        target_last_row = find_last_data_row(target_ws)
        target_row = target_last_row + 1
        clone_row_format(target_ws, target_last_row, target_row, 13)
        write_inventory_row(target_ws, target_row, updated)
        source_ws.delete_rows(current_row, 1)

    wb.save(filepath)
    return {"success": True, "rowsUpdated": 1}


def delete_inventory(wb, data, filepath):
    """Delete inventory rows from their workbook sheet before SQLite removes them."""
    items = data.get("items", [])
    deployments = data.get("deployments", [])
    if not items:
        return {"success": True, "rowsDeleted": 0}

    matches = []
    used_rows = set()
    for item in items:
        match = find_inventory_row(wb, item, minimum_score=6, early_return_score=8, used_rows=used_rows)
        if not match:
            identity = item.get("partNumber") or item.get("description") or "inventory row"
            return {"success": False, "error": f"Could not locate inventory row in Excel for deletion: {identity}"}
        used_rows.add(match)
        matches.append(match)

    deployment_matches = []
    used_deployment_rows = set()
    for deployment in deployments:
        match = find_deployment_row(wb, deployment, used_rows=used_deployment_rows)
        if not match:
            identity = deployment.get("partNumber") or deployment.get("description") or "deployment row"
            return {"success": False, "error": f"Could not locate linked deployment row in Excel for deletion: {identity}"}
        used_deployment_rows.add(match)
        deployment_matches.append(match)

    for sheet_name, row in sorted(deployment_matches, key=lambda value: (value[0], value[1]), reverse=True):
        wb[sheet_name].delete_rows(row, 1)

    for sheet_name, row in sorted(matches, key=lambda value: (value[0], value[1]), reverse=True):
        wb[sheet_name].delete_rows(row, 1)

    wb.save(filepath)
    return {"success": True, "rowsDeleted": len(matches), "deploymentRowsDeleted": len(deployment_matches)}


def deployment_match_score(ws, row, deployment):
    score = 0
    checks = [
        (1, normalized_text(deployment.get("itemType", ""))),
        (2, normalized_text(deployment.get("description", ""))),
        (3, normalized_text(deployment.get("manufacturer", ""))),
        (4, normalized_text(deployment.get("partNumber", ""))),
        (5, normalized_quantity(deployment.get("qtyDeployed", ""))),
        (6, normalized_text(deployment.get("deployedTo", ""))),
        (7, normalized_text(deployment.get("deployedBy", ""))),
        (8, normalized_date(deployment.get("deployedDate", ""))),
        (9, normalized_text(deployment.get("deployedLocation", ""))),
    ]

    for column, expected in checks:
        actual_value = ws.cell(row=row, column=column).value
        if column == 5:
            actual = normalized_quantity(actual_value)
        elif column == 8:
            actual = normalized_date(actual_value)
        else:
            actual = normalized_text(actual_value)
        if expected and actual == expected:
            score += 1

    return score


def find_deployment_row(wb, deployment, used_rows=None):
    if used_rows is None:
        used_rows = set()
    sheet_name = "Items Deployed"
    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]
    last_row = find_last_data_row(ws)
    best_match = None
    best_score = 0

    for row in range(2, last_row + 1):
        if (sheet_name, row) in used_rows:
            continue
        if not has_row_content(ws, row, [1, 2, 3, 4, 6, 7, 9]):
            continue
        score = deployment_match_score(ws, row, deployment)
        if score > best_score:
            best_match = (sheet_name, row)
            best_score = score
        if score >= 6:
            return (sheet_name, row)

    return best_match if best_score >= 5 else None


def delete_deployed(wb, data, filepath):
    """Delete deployment rows from the workbook so auto-sync cannot recreate them."""
    deployments = data.get("deployments", [])
    if not deployments:
        return {"success": True, "rowsDeleted": 0}

    matches = []
    used_rows = set()
    for deployment in deployments:
        match = find_deployment_row(wb, deployment, used_rows=used_rows)
        if not match:
            identity = deployment.get("partNumber") or deployment.get("description") or "deployment row"
            return {"success": False, "error": f"Could not locate deployment row in Excel for deletion: {identity}"}
        used_rows.add(match)
        matches.append(match)

    for sheet_name, row in sorted(matches, key=lambda value: (value[0], value[1]), reverse=True):
        wb[sheet_name].delete_rows(row, 1)

    wb.save(filepath)
    return {"success": True, "rowsDeleted": len(matches)}


def append_deployed(wb, data, filepath):
    """Append new rows to the Items Deployed sheet."""
    deployments = data.get("deployments", [])

    if not deployments:
        return {"success": True, "rowsAdded": 0}

    sheet_name = "Items Deployed"
    if sheet_name not in wb.sheetnames:
        return {"success": False, "error": f"Sheet '{sheet_name}' not found in workbook. Available sheets: {wb.sheetnames}"}

    ws = wb[sheet_name]
    last_row = find_last_data_row(ws)
    rows_added = 0

    for dep in deployments:
        row = last_row + 1 + rows_added
        clone_row_format(ws, last_row, row, 10)

        ws.cell(row=row, column=1, value=dep.get("itemType", ""))
        ws.cell(row=row, column=2, value=dep.get("description", ""))
        ws.cell(row=row, column=3, value=dep.get("manufacturer", ""))
        ws.cell(row=row, column=4, value=dep.get("partNumber", ""))

        # Qty deployed as number
        qty = dep.get("qtyDeployed")
        if qty not in (None, ""):
            ws.cell(row=row, column=5, value=parse_int(qty))

        ws.cell(row=row, column=6, value=dep.get("deployedTo", ""))
        ws.cell(row=row, column=7, value=dep.get("deployedBy", ""))

        # Deployed date as datetime
        deployed_date = parse_date(dep.get("deployedDate", ""))
        cell_h = ws.cell(row=row, column=8, value=deployed_date)
        if isinstance(deployed_date, datetime):
            cell_h.number_format = "MM/DD/YYYY"

        ws.cell(row=row, column=9, value=dep.get("deployedLocation", ""))
        ws.cell(row=row, column=10, value=dep.get("notes", ""))

        rows_added += 1

    wb.save(filepath)
    return {"success": True, "rowsAdded": rows_added}


def update_remaining(wb, data, filepath):
    """Update REMAINING INVENTORY (column L) using part number + PO + budget sheet when available."""
    updates = data if isinstance(data, list) else data.get("items", data.get("updates", []))

    if not updates:
        return {"success": True, "rowsUpdated": 0}

    lookup = {}
    part_only_lookup = {}
    for entry in updates:
        pn = str(entry.get("partNumber", "")).strip().casefold()
        po = str(entry.get("poNumber", "")).strip().casefold()
        budget_type = str(entry.get("budgetType", "Capital")).strip()
        sheet = "OpEx" if budget_type == "OpEx" else "Inventory"
        remaining = parse_int(entry.get("remaining", 0))
        if pn:
            lookup[(sheet, pn, po)] = remaining
            if not po:
                part_only_lookup[(sheet, pn)] = remaining

    rows_updated = 0

    for sheet_name in ["Inventory", "OpEx"]:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        last_row = find_last_data_row(ws)

        for row in range(2, last_row + 1):
            part_cell = ws.cell(row=row, column=4).value
            if part_cell is None:
                continue
            part_number = str(part_cell).strip().casefold()
            po_number = str(ws.cell(row=row, column=11).value or "").strip().casefold()
            key = (sheet_name, part_number, po_number)
            fallback_key = (sheet_name, part_number)
            if key in lookup:
                ws.cell(row=row, column=12, value=lookup[key])
                rows_updated += 1
            elif not po_number and fallback_key in part_only_lookup:
                ws.cell(row=row, column=12, value=part_only_lookup[fallback_key])
                rows_updated += 1

    wb.save(filepath)
    return {"success": True, "rowsUpdated": rows_updated}


def read_inventory(wb, data, filepath):
    """Read all rows from the Inventory and/or OpEx sheets and return as JSON."""
    sheets_to_read = data.get("sheets", ["Inventory", "OpEx"])
    all_items = []

    for sheet_name in sheets_to_read:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        last_row = find_last_data_row(ws)
        budget_type = "OpEx" if sheet_name == "OpEx" else "Capital"

        for row in range(2, last_row + 1):  # skip header
            if not has_row_content(ws, row, [1, 2, 3, 4, 6, 11]):
                continue  # skip empty rows
            item_type = ws.cell(row=row, column=1).value

            # Parse purchase date
            date_val = ws.cell(row=row, column=5).value
            purchase_date = ""
            if isinstance(date_val, datetime):
                purchase_date = date_val.strftime("%m/%d/%Y")
            elif isinstance(date_val, date):
                purchase_date = date_val.strftime("%m/%d/%Y")
            elif date_val:
                purchase_date = str(date_val)

            # Parse qty received string like "20/20"
            qty_recv_str = str(ws.cell(row=row, column=10).value or "").strip()
            qty_received = 0
            if "/" in qty_recv_str:
                qty_received = parse_int(qty_recv_str.split("/")[0], default=0)
            elif qty_recv_str:
                qty_received = parse_int(qty_recv_str, default=0)

            # Unit cost
            unit_cost = ws.cell(row=row, column=7).value
            unit_cost = parse_float(unit_cost, default=0)

            # Quantity
            quantity = ws.cell(row=row, column=8).value
            quantity = parse_int(quantity, default=0)

            # PO number as string
            po_val = ws.cell(row=row, column=11).value
            po_number = str(int(po_val)) if isinstance(po_val, float) else (str(po_val) if po_val else "")

            item = {
                "itemType": str(item_type).strip(),
                "description": str(ws.cell(row=row, column=2).value or "").strip(),
                "manufacturer": str(ws.cell(row=row, column=3).value or "").strip(),
                "partNumber": str(ws.cell(row=row, column=4).value or "").strip(),
                "purchaseDate": purchase_date,
                "vendor": str(ws.cell(row=row, column=6).value or "").strip(),
                "unitCost": unit_cost,
                "quantity": quantity,
                "qtyReceived": qty_received,
                "poNumber": po_number,
                "notes": str(ws.cell(row=row, column=13).value or "").strip(),
                "budgetType": budget_type,
            }
            all_items.append(item)

    return {"success": True, "items": all_items, "count": len(all_items)}


def read_deployed(wb, data, filepath):
    """Read all rows from the Items Deployed sheet and return as JSON."""
    sheet_name = "Items Deployed"
    if sheet_name not in wb.sheetnames:
        return {"success": False, "error": f"Sheet '{sheet_name}' not found"}

    ws = wb[sheet_name]
    last_row = find_last_data_row(ws)
    deployments = []

    for row in range(2, last_row + 1):
        if not has_row_content(ws, row, [1, 2, 3, 4, 6, 7, 9]):
            continue
        item_type = ws.cell(row=row, column=1).value

        date_val = ws.cell(row=row, column=8).value
        deployed_date = ""
        if isinstance(date_val, datetime):
            deployed_date = date_val.strftime("%m/%d/%Y")
        elif isinstance(date_val, date):
            deployed_date = date_val.strftime("%m/%d/%Y")
        elif date_val:
            deployed_date = str(date_val)

        qty = parse_int(ws.cell(row=row, column=5).value, default=1)

        dep = {
            "itemType": str(item_type).strip(),
            "description": str(ws.cell(row=row, column=2).value or "").strip(),
            "manufacturer": str(ws.cell(row=row, column=3).value or "").strip(),
            "partNumber": str(ws.cell(row=row, column=4).value or "").strip(),
            "qtyDeployed": qty,
            "deployedTo": str(ws.cell(row=row, column=6).value or "").strip(),
            "deployedBy": str(ws.cell(row=row, column=7).value or "").strip(),
            "deployedDate": deployed_date,
            "deployedLocation": str(ws.cell(row=row, column=9).value or "").strip(),
            "notes": str(ws.cell(row=row, column=10).value or "").strip(),
        }
        deployments.append(dep)

    return {"success": True, "deployments": deployments, "count": len(deployments)}


def main():
    if len(sys.argv) < 3:
        print(json.dumps({"success": False, "error": "Usage: excel_sync.py <command> <excel_path>"}))
        sys.exit(1)

    command = sys.argv[1]
    filepath = sys.argv[2]

    # Read JSON from stdin
    try:
        raw_input = sys.stdin.read()
        data = json.loads(raw_input) if raw_input.strip() else {}
    except json.JSONDecodeError as e:
        print(json.dumps({"success": False, "error": f"Invalid JSON input: {str(e)}"}))
        sys.exit(1)

    # Load workbook
    try:
        wb = openpyxl.load_workbook(filepath)
    except FileNotFoundError:
        print(json.dumps({"success": False, "error": f"Excel file not found: {filepath}"}))
        sys.exit(1)
    except Exception as e:
        print(json.dumps({"success": False, "error": f"Failed to open workbook: {str(e)}"}))
        sys.exit(1)

    # Dispatch command
    try:
        if command == "append-inventory":
            result = append_inventory(wb, data, filepath)
        elif command == "append-deployed":
            result = append_deployed(wb, data, filepath)
        elif command == "update-inventory":
            result = update_inventory(wb, data, filepath)
        elif command == "delete-inventory":
            result = delete_inventory(wb, data, filepath)
        elif command == "delete-deployed":
            result = delete_deployed(wb, data, filepath)
        elif command == "update-remaining":
            result = update_remaining(wb, data, filepath)
        elif command == "read-inventory":
            result = read_inventory(wb, data, filepath)
        elif command == "read-deployed":
            result = read_deployed(wb, data, filepath)
        else:
            result = {"success": False, "error": f"Unknown command: {command}"}
    except Exception as e:
        result = {"success": False, "error": str(e)}

    print(json.dumps(result))


if __name__ == "__main__":
    main()
